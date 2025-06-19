import numpy as np
import openpyxl

class EpochDataSaver:
	def __init__(self):
		self.goal = []
		self.len = []
		self.time = []
		self.ret = []
		self.dr = []
		self.avg_ret = []
		self.suc_rate = []
		self.suc_idx = []
		self.suc_len = []
		self.suc_time = []
		self.suc_ret = []
		self.suc_avg_ret = []

	def push_data(self, g, r, l, t, dr):
		self.goal.append(g)
		self.len.append(l)
		self.time.append(t)
		self.ret.append(r)
		self.dr.append(dr)
		self.avg_ret.append(sum(self.ret)/len(self.ret))
		if dr == 'success':
			self.suc_idx.append(len(self.ret))
			self.suc_len.append(l)
			self.suc_time.append(t)
			self.suc_ret.append(r)
			self.suc_avg_ret.append(sum(self.suc_ret)/len(self.suc_ret))
		self.suc_rate.append(round(100*len(self.suc_ret)/len(self.ret), 4))

	def save_data_file(self, filename):
		filepath = '/home/zsw/catkin_ws/src/turtlebot3_drl_pp/results/'+filename+'.npz'
		np.savez(
			filepath,
			goal = self.goal,
			len = self.len,
			time = self.time,
			ret = self.ret,
			dr = self.dr,
			avg_ret = self.avg_ret,
			suc_rate = self.suc_rate,
			suc_idx = self.suc_idx,
			suc_len = self.suc_len,
			suc_time = self.suc_time,
			suc_ret = self.suc_ret,
			suc_avg_ret = self.suc_avg_ret
		)

	def load_data_file(self, filename):
		filepath = '/home/zsw/catkin_ws/src/turtlebot3_drl_pp/results/'+filename+'.npz'
		data = np.load(filepath)
		self.goal = data['goal'].tolist()
		self.len = data['len'].tolist()
		self.time = data['time'].tolist()
		self.ret = data['ret'].tolist()
		self.dr = data['dr'].tolist()
		self.avg_ret = data['avg_ret'].tolist()
		self.suc_rate = data['suc_rate'].tolist()
		self.suc_idx = data['suc_idx'].tolist()
		self.suc_len = data['suc_len'].tolist()
		self.suc_time = data['suc_time'].tolist()
		self.suc_ret = data['suc_ret'].tolist()
		self.suc_avg_ret = data['suc_avg_ret'].tolist()

	def save_data_excel(self, filename):
		filepath = '/home/zsw/catkin_ws/src/turtlebot3_drl_pp/results/'+filename+'.xlsx'
		workbook = openpyxl.Workbook()
		
		worksheet1 = workbook.active
		worksheet1.title = 'epoch_data'
		tabular1 = ['epoch', 'goal', 'len', 'time', 'ret', 'done_reason', 'avg_ret', 'success_rate']
		for i in range(len(tabular1)):
			worksheet1.cell(1, i+1, tabular1[i])
		for i in range(len(self.ret)):
			worksheet1.cell(i+2, 1, i+1)
			worksheet1.cell(i+2, 2, self.goal[i])
			worksheet1.cell(i+2, 3, self.len[i])
			worksheet1.cell(i+2, 4, self.time[i])
			worksheet1.cell(i+2, 5, self.ret[i])
			worksheet1.cell(i+2, 6, self.dr[i])
			worksheet1.cell(i+2, 7, self.avg_ret[i])
			worksheet1.cell(i+2, 8, self.suc_rate[i])

		worksheet2 = workbook.create_sheet()
		worksheet2.title = 'suc_epoch_data'
		tabular2 = ['suc_epoch', 'suc_len', 'suc_time', 'suc_ret', 'suc_avg_ret']
		for i in range(len(tabular2)):
			worksheet2.cell(1, i+1, tabular2[i])
		for i in range(len(self.suc_idx)):
			worksheet2.cell(i+2, 1, self.suc_idx[i])
			worksheet2.cell(i+2, 2, self.suc_len[i])
			worksheet2.cell(i+2, 3, self.suc_time[i])
			worksheet2.cell(i+2, 4, self.suc_ret[i])
			worksheet2.cell(i+2, 5, self.suc_avg_ret[i])
		
		workbook.save(filepath)

class PathsSaver():
	def __init__(self):
		self.path = []

	def push_path(self):
		self.path.append([])

	def push_point(self, data):
		self.path[-1].append(data)
	
	def save_paths(self, filename):
		filepath = '/home/zsw/catkin_ws/src/turtlebot3_drl_pp/results/'+filename+'(path).xlsx'
		workbook = openpyxl.Workbook()
		worksheet = workbook.active
		worksheet.title = 'test_paths'
		for i in range(len(self.path)):
			worksheet.cell(1, 2*i+1, 'path_'+str(i+1)+'_x')
			worksheet.cell(1, 2*i+2, 'path_'+str(i+1)+'_y')
			for j in range(len(self.path[i])):
				worksheet.cell(j+2, 2*i+1, self.path[i][j][0])
				worksheet.cell(j+2, 2*i+2, self.path[i][j][1])
		workbook.save(filepath)
		